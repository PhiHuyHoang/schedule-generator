from icalendar import Calendar, Event
from pytz import UTC # timezone
from datetime import  datetime,timedelta
import openpyxl

cal = Calendar()
cal.add('prodid', '-//My calendar product//mxm.dk//')
cal.add('version', '2.0')

def make_ics(work,i):
    event = Event()
    event.add('summary', work[0])
    event.add('dtstart', datetime(work[1].year,work[1].month,work[1].day,work[2].hour-2,work[2].minute,0,tzinfo=UTC))
    event.add('dtend', datetime(work[3].year,work[3].month,work[3].day,work[4].hour-2,work[4].minute,0,tzinfo=UTC))
    event.add('dtstamp', datetime.now())
    event.add('location', work[7])
    event['uid'] = str(i)+'@mxm.dk'
    event.add('priority', 5)

    cal.add_component(event)


wb = openpyxl.load_workbook('hoang-neptun.xlsx')
sheet = wb['hoang-neptun']
for i in range(2,95):
    work = []
    for j in range(1, sheet.max_column + 1):
        work.append(sheet.cell(row = i,column = j).value)
    try:
        make_ics(work,i)
    except Exception as e:
        print(e)
f = open('Personal-schedule.ics', 'wb')
f.write(cal.to_ical())
f.close()

# def all_day_left(year):
#    d = datetime.now().date()                # January 1st
#    d += timedelta(days = 1)  # First Sunday
#    while d.year == year:
#       yield d
#       d += timedelta(days = 1)
#
# list_of_day_left = all_day_left(datetime.now().year)
#
# for day in list_of_day_left:
#     if day.weekday() in [2]:
#         print(day)